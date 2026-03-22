import os
from openpyxl import load_workbook

DATA_PATH = os.path.join(os.path.dirname(__file__), '..', 'data.xlsx')

PROJECTS = {
    'מנרב סנטר': {'start_row': 6, 'end_row': 17, 'summary_row': 18},
    'קריית מנרב': {'start_row': 22, 'end_row': 33, 'summary_row': 34},
    'הדסה': {'start_row': 38, 'end_row': 49, 'summary_row': 50},
    'עמיגור באר שבע': {'start_row': 57, 'end_row': 68, 'summary_row': 69},
}

MONTH_TO_COL = {9: 5, 10: 7, 11: 9, 12: 11, 1: 13, 2: 15, 3: 17, 4: 19, 5: 21, 6: 23, 7: 25, 8: 27, 9: 29}

CASHFLOW_ROWS = {
    'קריית מנרב': 4,
    'מנרב סנטר': 5,
    'הדסה': 6,
    'עמיגור באר שבע': 7,
}

PROJECT_META = {
    'מנרב סנטר': {'manager': 'אתי', 'area': 'מסחרי פרטי', 'axis': 'FM', 'priority_id': 'P-1001'},
    'קריית מנרב': {'manager': 'אתי', 'area': 'מסחרי פרטי', 'axis': 'FM', 'priority_id': 'P-1002'},
    'הדסה': {'manager': 'אתי', 'area': 'מסחרי פרטי', 'axis': 'FM', 'priority_id': 'P-1003'},
    'עמיגור באר שבע': {'manager': 'אלון', 'area': 'פרוייקטים', 'axis': 'FM', 'priority_id': 'P-2001'},
}

# Monthly expense breakdown components per project (from Excel side columns)
EXPENSE_BREAKDOWN = {
    'מנרב סנטר': {
        'op_components': [
            {'name': 'ניקיון ותחזוקה', 'amount': 100},
        ],
        'salary_components': [
            {'name': 'שכר מנרב אחזקות', 'amount': 100},
            {'name': 'העמסת שכר עובד 1', 'amount': 16.67},
            {'name': 'העמסת שכר עובד 2', 'amount': 1.67},
        ],
    },
    'קריית מנרב': {
        'op_components': [
            {'name': 'ניקיון ותחזוקה', 'amount': 255},
            {'name': 'פער ספק ניקיון', 'amount': 45},
            {'name': 'הוצאות שוטפות', 'amount': 98.84},
        ],
        'salary_components': [
            {'name': 'שכר מנרב אחזקות', 'amount': 250},
            {'name': 'העמסת שכר עובד 1', 'amount': 16.67},
            {'name': 'העמסת שכר עובד 2', 'amount': 1.67},
        ],
    },
    'הדסה': {
        'op_components': [
            {'name': 'ניקיון ותחזוקה', 'amount': 100},
            {'name': 'הוצאות שוטפות', 'amount': 29.8},
        ],
        'salary_components': [
            {'name': 'שכר מנרב אחזקות', 'amount': 70},
            {'name': 'העמסת שכר עובד 1 (שליש)', 'amount': 16.67},
            {'name': 'העמסת שכר עובד 2', 'amount': 1.67},
        ],
    },
    'עמיגור באר שבע': {
        'op_components': [
            {'name': 'עבודות פנים', 'amount': None},
            {'name': 'עבודות חוץ', 'amount': None},
        ],
        'salary_components': [],
        'milestones': [
            {'name': 'פעימה 1', 'revenue': 70, 'expense': 0, 'month': 'פברואר'},
            {'name': 'פעימה 2', 'revenue': 150, 'expense': 0, 'month': 'פברואר'},
            {'name': 'פעימה 3', 'revenue': 450, 'expense': 540, 'month': 'מרץ'},
            {'name': 'פעימה 4', 'revenue': 336.36, 'expense': 302.73, 'month': 'מאי'},
        ],
    },
}


def _val(v):
    if v is None:
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def load_pnl(project_name=None):
    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb['פרוייקטים מנרב IFM']
    result = {}

    targets = {project_name: PROJECTS[project_name]} if project_name else PROJECTS

    for name, info in targets.items():
        months = []
        for r in range(info['start_row'], info['end_row'] + 1):
            month_num = ws.cell(row=r, column=2).value
            revenue = _val(ws.cell(row=r, column=3).value)
            op_expenses = _val(ws.cell(row=r, column=4).value)
            gross_profit = _val(ws.cell(row=r, column=5).value)
            salary_expenses = _val(ws.cell(row=r, column=6).value)
            operating_profit = _val(ws.cell(row=r, column=7).value)
            notes = ws.cell(row=r, column=8).value or ''

            margin = round((operating_profit / revenue) * 100, 1) if revenue > 0 else None

            months.append({
                'month': int(month_num) if month_num else 0,
                'revenue': round(revenue, 2),
                'op_expenses': round(op_expenses, 2),
                'gross_profit': round(gross_profit, 2),
                'salary_expenses': round(salary_expenses, 2),
                'operating_profit': round(operating_profit, 2),
                'margin': margin,
                'margin_alert': margin is not None and margin < 20,
                'notes': str(notes),
            })

        sr = info['summary_row']
        summary = {
            'total_revenue': round(_val(ws.cell(row=sr, column=3).value), 2),
            'total_op_expenses': round(_val(ws.cell(row=sr, column=4).value), 2),
            'total_gross_profit': round(_val(ws.cell(row=sr, column=5).value), 2),
            'total_salary_expenses': round(_val(ws.cell(row=sr, column=6).value), 2),
            'total_operating_profit': round(_val(ws.cell(row=sr, column=7).value), 2),
        }
        total_rev = summary['total_revenue']
        summary['margin'] = round((summary['total_operating_profit'] / total_rev) * 100, 1) if total_rev > 0 else None

        meta = PROJECT_META.get(name, {})
        expense_breakdown = EXPENSE_BREAKDOWN.get(name, {})
        result[name] = {'months': months, 'summary': summary, 'meta': meta, 'expense_breakdown': expense_breakdown}

    wb.close()
    return result


def load_project_cashflow(project_name):
    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb['ריכוז']
    row_num = CASHFLOW_ROWS.get(project_name)
    if not row_num:
        wb.close()
        return None

    month_labels = ['9/2025', '10/2025', '11/2025', '12/2025',
                    '1/2026', '2/2026', '3/2026', '4/2026', '5/2026',
                    '6/2026', '7/2026', '8/2026', '9/2026']
    col_map = [5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29]

    data = []
    cumulative = 0
    for ci, col in enumerate(col_map):
        exp = _val(ws.cell(row=row_num, column=col).value)
        rev = _val(ws.cell(row=row_num, column=col + 1).value)
        net = rev - exp
        cumulative += net
        data.append({
            'month': month_labels[ci],
            'expenses': round(exp, 2),
            'revenue': round(rev, 2),
            'net': round(net, 2),
            'cumulative': round(cumulative, 2),
        })

    wb.close()
    return {'month_labels': month_labels, 'data': data}


def load_cashflow():
    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb['ריכוז']

    month_labels = ['9/2025', '10/2025', '11/2025', '12/2025',
                    '1/2026', '2/2026', '3/2026', '4/2026', '5/2026',
                    '6/2026', '7/2026', '8/2026', '9/2026']
    month_keys = [9, 10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    # Column mapping: expense col, revenue col = col, col+1
    col_map = [5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29]

    projects = {}
    for pname, row_num in CASHFLOW_ROWS.items():
        data = []
        for ci, col in enumerate(col_map):
            exp = _val(ws.cell(row=row_num, column=col).value)
            rev = _val(ws.cell(row=row_num, column=col + 1).value)
            data.append({
                'month': month_labels[ci],
                'expenses': round(exp, 2),
                'revenue': round(rev, 2),
                'profit': round(rev - exp, 2),
            })
        projects[pname] = data

    # Totals row (12)
    totals_row = 12
    totals = []
    for ci, col in enumerate(col_map):
        exp = _val(ws.cell(row=totals_row, column=col).value)
        rev = _val(ws.cell(row=totals_row, column=col + 1).value)
        totals.append({
            'month': month_labels[ci],
            'expenses': round(exp, 2),
            'revenue': round(rev, 2),
        })

    # Monthly net (row 14)
    monthly_net = []
    for ci, col in enumerate(col_map):
        v = _val(ws.cell(row=14, column=col).value)
        monthly_net.append({'month': month_labels[ci], 'value': round(v, 2)})

    # Cumulative (row 15)
    cumulative = []
    for ci, col in enumerate(col_map):
        v = _val(ws.cell(row=15, column=col).value)
        cumulative.append({'month': month_labels[ci], 'value': round(v, 2)})

    wb.close()
    return {
        'month_labels': month_labels,
        'projects': projects,
        'totals': totals,
        'monthly_net': monthly_net,
        'cumulative': cumulative,
    }


def load_dashboard_kpis():
    pnl = load_pnl()
    total_revenue = sum(p['summary']['total_revenue'] for p in pnl.values())
    total_expenses = sum(p['summary']['total_op_expenses'] + p['summary']['total_salary_expenses'] for p in pnl.values())
    total_op_profit = sum(p['summary']['total_operating_profit'] for p in pnl.values())
    margin = round((total_op_profit / total_revenue) * 100, 1) if total_revenue > 0 else None

    cf = load_cashflow()
    cash_position = cf['cumulative'][-1]['value'] if cf['cumulative'] else 0

    return {
        'total_revenue': round(total_revenue, 2),
        'total_expenses': round(total_expenses, 2),
        'total_operating_profit': round(total_op_profit, 2),
        'margin': margin,
        'cash_position': round(cash_position, 2),
        'project_summaries': {name: data['summary'] for name, data in pnl.items()},
    }


# --- Unified Data Layer ---

PROJECTS_DATA_FILE = os.path.join(os.path.dirname(DATA_PATH), 'projects_data.json')


def _load_form_data():
    if not os.path.exists(PROJECTS_DATA_FILE):
        return {}
    import json
    with open(PROJECTS_DATA_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def _save_form_data(data):
    import json
    with open(PROJECTS_DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _get_shotef_offset(payment_terms):
    """Calculate payment delay in months from payment terms list (weighted average, legacy)."""
    if not payment_terms:
        return 0
    weighted = 0
    for t in payment_terms:
        pct = t.get('percent', 0) / 100
        term_type = t.get('type', '')
        if 'שוטף+90' in term_type:
            weighted += pct * 3
        elif 'שוטף+60' in term_type:
            weighted += pct * 2
        elif 'שוטף+45' in term_type:
            weighted += pct * 1.5
        elif 'שוטף+30' in term_type:
            weighted += pct * 1
        # מזומן and מקדמה = 0 offset
    return round(weighted)


def _shotef_offset_single(term_type):
    """Get payment delay in months for a single payment term type."""
    if not term_type or term_type in ('מזומן', 'מקדמה'):
        return 0
    if 'שוטף+90' in term_type:
        return 3
    if 'שוטף+60' in term_type:
        return 2
    if 'שוטף+45' in term_type:
        return 2
    if 'שוטף+30' in term_type:
        return 1
    return 0


def _calc_revenue_for_month(form_data, target_cal_month):
    """Calculate revenue arriving in a specific calendar month using per-term logic."""
    total_revenue = form_data.get('total_revenue', 0) or 0
    if not total_revenue:
        return 0
    forecast = form_data.get('revenue_forecast', {})
    payment_terms = form_data.get('revenue_payment_terms', [])
    if not payment_terms:
        payment_terms = [{'type': 'מזומן', 'percent': 100}]

    total = 0
    for invoice_m in range(1, 13):
        pct = forecast.get(str(invoice_m), 0) or 0
        if not pct:
            continue
        invoice_amount = total_revenue * pct / 100
        for term in payment_terms:
            term_pct = term.get('percent', 0) or 0
            if not term_pct:
                continue
            offset = _shotef_offset_single(term.get('type', ''))
            arrival_m = invoice_m + offset
            if arrival_m == target_cal_month:
                total += round(invoice_amount * term_pct / 100)
    return total


def form_to_pnl(form_data, current_month=None):
    """Convert form data to P&L format. Uses actuals for past months, forecast for future."""
    import datetime as dt
    if current_month is None:
        current_month = dt.datetime.now().month

    actuals = form_data.get('actuals', {})
    months = []

    for m in range(1, 13):
        actual = actuals.get(str(m), {})
        has_actual = actual and (actual.get('revenue', 0) > 0 or actual.get('op_expenses', 0) > 0)

        if has_actual and m <= current_month:
            # Use actuals
            revenue = actual.get('revenue', 0)
            op_expenses = actual.get('op_expenses', 0)
            salary_expenses = actual.get('salary_expenses', 0)
        else:
            # Use forecast
            revenue = 0
            if form_data.get('total_revenue') and form_data.get('revenue_forecast', {}).get(str(m)):
                revenue = round(form_data['total_revenue'] * form_data['revenue_forecast'][str(m)] / 100)

            op_expenses = 0
            for cat in ['manpower', 'equipment', 'insurance', 'consultants', 'financing', 'other']:
                for line in form_data.get(f'expense_lines_{cat}', []):
                    start = line.get('start_month', 1)
                    end = line.get('end_month', 12)
                    if start <= m <= end:
                        op_expenses += line.get('monthly_amount', 0) or 0
            for sub in form_data.get('subcontractors', []):
                op_expenses += sub.get('monthly_amount', 0) or 0

            salary_expenses = 0

        operating_profit = revenue - op_expenses - salary_expenses
        margin = round((operating_profit / revenue) * 100, 1) if revenue > 0 else None

        months.append({
            'month': m,
            'revenue': round(revenue, 2),
            'op_expenses': round(op_expenses, 2),
            'gross_profit': round(revenue - op_expenses, 2),
            'salary_expenses': round(salary_expenses, 2),
            'operating_profit': round(operating_profit, 2),
            'margin': margin,
            'margin_alert': margin is not None and margin < 20,
            'notes': actual.get('notes', ''),
            'is_actual': has_actual and m <= current_month,
        })

    total_rev = sum(m['revenue'] for m in months)
    total_op = sum(m['op_expenses'] for m in months)
    total_sal = sum(m['salary_expenses'] for m in months)
    total_profit = total_rev - total_op - total_sal

    summary = {
        'total_revenue': round(total_rev, 2),
        'total_op_expenses': round(total_op, 2),
        'total_gross_profit': round(total_rev - total_op, 2),
        'total_salary_expenses': round(total_sal, 2),
        'total_operating_profit': round(total_profit, 2),
        'margin': round((total_profit / total_rev) * 100, 1) if total_rev > 0 else None,
    }

    meta = {
        'manager': form_data.get('manager', ''),
        'area': form_data.get('area', ''),
        'axis': form_data.get('axis', ''),
        'priority_id': form_data.get('priority_id', ''),
    }

    return {'months': months, 'summary': summary, 'meta': meta, 'expense_breakdown': {}}


def form_to_cashflow(form_data):
    """Compute cashflow from form data including payment term delays."""
    month_labels = ['9/2025', '10/2025', '11/2025', '12/2025',
                    '1/2026', '2/2026', '3/2026', '4/2026', '5/2026',
                    '6/2026', '7/2026', '8/2026', '9/2026']
    # Map month_labels index to calendar month (9,10,11,12,1,2,...,9)
    cal_months = [9, 10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]

    actuals = form_data.get('actuals', {})

    data = []
    cumulative = 0

    for ci, cal_m in enumerate(cal_months):
        actual = actuals.get(str(cal_m), {})
        has_actual = actual and (actual.get('revenue', 0) > 0 or actual.get('op_expenses', 0) > 0)

        if has_actual:
            rev = actual.get('revenue', 0)
            exp = actual.get('op_expenses', 0) + actual.get('salary_expenses', 0)
        else:
            # Forecast revenue with per-term payment delay
            rev = _calc_revenue_for_month(form_data, cal_m)

            # Forecast expenses
            exp = 0
            for cat in ['manpower', 'equipment', 'insurance', 'consultants', 'financing', 'other']:
                for line in form_data.get(f'expense_lines_{cat}', []):
                    start = line.get('start_month', 1)
                    end = line.get('end_month', 12)
                    if start <= cal_m <= end:
                        exp += line.get('monthly_amount', 0) or 0
            for sub in form_data.get('subcontractors', []):
                exp += sub.get('monthly_amount', 0) or 0

        net = rev - exp
        cumulative += net
        data.append({
            'month': month_labels[ci],
            'expenses': round(exp, 2),
            'revenue': round(rev, 2),
            'net': round(net, 2),
            'cumulative': round(cumulative, 2),
        })

    return {'month_labels': month_labels, 'data': data}


def load_unified_projects():
    """Load all projects from both Excel and form. Form data takes priority."""
    form_data = _load_form_data()

    # Start with Excel projects
    try:
        excel_pnl = load_pnl()
    except Exception:
        excel_pnl = {}

    result = {}

    # Add Excel projects (form override if exists)
    for name, pnl_data in excel_pnl.items():
        if name in form_data:
            # Form takes priority
            result[name] = {
                'pnl': form_to_pnl(form_data[name]),
                'source': form_data[name].get('source', 'form'),
                'status': form_data[name].get('status', 'active'),
                'last_updated': form_data[name].get('last_updated', ''),
            }
        else:
            result[name] = {
                'pnl': pnl_data,
                'source': 'excel',
                'status': 'active',
                'last_updated': '',
            }

    # Add form-only projects
    for name, fdata in form_data.items():
        if name not in result:
            result[name] = {
                'pnl': form_to_pnl(fdata),
                'source': fdata.get('source', 'form'),
                'status': fdata.get('status', 'active'),
                'last_updated': fdata.get('last_updated', ''),
            }

    return result


def load_unified_dashboard():
    """Dashboard KPIs from all projects (unified)."""
    unified = load_unified_projects()

    total_revenue = 0
    total_expenses = 0
    total_op_profit = 0
    project_summaries = {}

    for name, data in unified.items():
        s = data['pnl']['summary']
        total_revenue += s.get('total_revenue', 0)
        total_expenses += s.get('total_op_expenses', 0) + s.get('total_salary_expenses', 0)
        total_op_profit += s.get('total_operating_profit', 0)
        project_summaries[name] = {
            **s,
            'meta': data['pnl'].get('meta', {}),
            'source': data['source'],
            'status': data['status'],
        }

    margin = round((total_op_profit / total_revenue) * 100, 1) if total_revenue > 0 else None

    # Try to get cash position from Excel if available
    try:
        cf = load_cashflow()
        cash_position = cf['cumulative'][-1]['value'] if cf['cumulative'] else 0
    except Exception:
        cash_position = 0

    return {
        'total_revenue': round(total_revenue, 2),
        'total_expenses': round(total_expenses, 2),
        'total_operating_profit': round(total_op_profit, 2),
        'margin': margin,
        'cash_position': round(cash_position, 2),
        'project_summaries': project_summaries,
    }


def import_excel_to_form(project_name):
    """Import an Excel project into form data structure."""
    if project_name not in PROJECTS:
        return None

    pnl = load_pnl(project_name)
    project_pnl = pnl[project_name]
    meta = PROJECT_META.get(project_name, {})
    breakdown = EXPENSE_BREAKDOWN.get(project_name, {})

    # Build revenue forecast from monthly data
    total_rev = project_pnl['summary']['total_revenue']
    revenue_forecast = {}
    for m_data in project_pnl['months']:
        m = m_data['month']
        if total_rev > 0:
            revenue_forecast[str(m)] = round(m_data['revenue'] / total_rev * 100, 1)
        else:
            revenue_forecast[str(m)] = 0

    # Build expense lines from breakdown
    expense_lines_other = []
    for comp in breakdown.get('op_components', []):
        if comp.get('amount'):
            expense_lines_other.append({
                'name': comp['name'],
                'monthly_amount': comp['amount'],
                'start_month': 1,
                'end_month': 12,
            })

    expense_lines_manpower = []
    for comp in breakdown.get('salary_components', []):
        if comp.get('amount'):
            expense_lines_manpower.append({
                'name': comp['name'],
                'monthly_amount': comp['amount'],
                'start_month': 1,
                'end_month': 12,
            })

    # Build actuals from Excel data
    actuals = {}
    for m_data in project_pnl['months']:
        m = m_data['month']
        actuals[str(m)] = {
            'revenue': m_data['revenue'],
            'op_expenses': m_data['op_expenses'],
            'salary_expenses': m_data['salary_expenses'],
            'notes': m_data.get('notes', ''),
        }

    import datetime as dt
    form_data = {
        'project_name': project_name,
        'priority_id': meta.get('priority_id', ''),
        'start_date': '',
        'description': f'פרויקט מיובא מ-Excel',
        'manager': meta.get('manager', ''),
        'client': '',
        'area': meta.get('area', ''),
        'axis': meta.get('axis', ''),
        'total_revenue': total_rev,
        'revenue_payment_terms': [{'type': 'שוטף+30', 'percent': 100}],
        'revenue_forecast': revenue_forecast,
        'total_budget': project_pnl['summary']['total_op_expenses'] + project_pnl['summary']['total_salary_expenses'],
        'subcontractors': [],
        'expense_lines_manpower': expense_lines_manpower,
        'expense_lines_equipment': [],
        'expense_lines_insurance': [],
        'expense_lines_consultants': [],
        'expense_lines_financing': [],
        'expense_lines_other': expense_lines_other,
        'actuals': actuals,
        'status': 'active',
        'source': 'excel-import',
        'last_updated': dt.datetime.now().isoformat(),
    }

    # Save to projects_data.json
    all_data = _load_form_data()
    all_data[project_name] = form_data
    _save_form_data(all_data)

    return form_data
