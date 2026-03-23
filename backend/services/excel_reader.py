from openpyxl import load_workbook
from config import DATA_PATH, PROJECTS, CASHFLOW_ROWS, PROJECT_META, EXPENSE_BREAKDOWN, MONTH_LABELS, COL_MAP


def _val(v):
    if v is None:
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def load_pnl(project_name=None):
    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb['פרוייקטים מנרב IFM']
    result = {}

    targets = {project_name: PROJECTS[project_name]} if project_name else PROJECTS

    for name, info in targets.items():
        months = []
        for r in range(info['start_row'], info['end_row'] + 1):
            month_num = ws.cell(row=r, column=2).value
            revenue = _val(ws.cell(row=r, column=3).value)
            op_expenses = _val(ws.cell(row=r, column=4).value)
            gross_profit = _val(ws.cell(row=r, column=5).value)
            salary_expenses = _val(ws.cell(row=r, column=6).value)
            operating_profit = _val(ws.cell(row=r, column=7).value)
            notes = ws.cell(row=r, column=8).value or ''

            margin = round((operating_profit / revenue) * 100, 1) if revenue > 0 else None

            months.append({
                'month': int(month_num) if month_num else 0,
                'revenue': round(revenue, 2),
                'op_expenses': round(op_expenses, 2),
                'gross_profit': round(gross_profit, 2),
                'salary_expenses': round(salary_expenses, 2),
                'operating_profit': round(operating_profit, 2),
                'margin': margin,
                'margin_alert': margin is not None and margin < 20,
                'notes': str(notes),
            })

        sr = info['summary_row']
        summary = {
            'total_revenue': round(_val(ws.cell(row=sr, column=3).value), 2),
            'total_op_expenses': round(_val(ws.cell(row=sr, column=4).value), 2),
            'total_gross_profit': round(_val(ws.cell(row=sr, column=5).value), 2),
            'total_salary_expenses': round(_val(ws.cell(row=sr, column=6).value), 2),
            'total_operating_profit': round(_val(ws.cell(row=sr, column=7).value), 2),
        }
        total_rev = summary['total_revenue']
        summary['margin'] = round((summary['total_operating_profit'] / total_rev) * 100, 1) if total_rev > 0 else None

        meta = PROJECT_META.get(name, {})
        expense_breakdown = EXPENSE_BREAKDOWN.get(name, {})
        result[name] = {'months': months, 'summary': summary, 'meta': meta, 'expense_breakdown': expense_breakdown}

    wb.close()
    return result


def load_project_cashflow(project_name):
    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb['ריכוז']
    row_num = CASHFLOW_ROWS.get(project_name)
    if not row_num:
        wb.close()
        return None

    data = []
    cumulative = 0
    for ci, col in enumerate(COL_MAP):
        exp = _val(ws.cell(row=row_num, column=col).value)
        rev = _val(ws.cell(row=row_num, column=col + 1).value)
        net = rev - exp
        cumulative += net
        data.append({
            'month': MONTH_LABELS[ci],
            'expenses': round(exp, 2),
            'revenue': round(rev, 2),
            'net': round(net, 2),
            'cumulative': round(cumulative, 2),
        })

    wb.close()
    return {'month_labels': MONTH_LABELS, 'data': data}


def load_cashflow():
    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb['ריכוז']

    projects = {}
    for pname, row_num in CASHFLOW_ROWS.items():
        data = []
        for ci, col in enumerate(COL_MAP):
            exp = _val(ws.cell(row=row_num, column=col).value)
            rev = _val(ws.cell(row=row_num, column=col + 1).value)
            data.append({
                'month': MONTH_LABELS[ci],
                'expenses': round(exp, 2),
                'revenue': round(rev, 2),
                'profit': round(rev - exp, 2),
            })
        projects[pname] = data

    totals_row = 12
    totals = []
    for ci, col in enumerate(COL_MAP):
        exp = _val(ws.cell(row=totals_row, column=col).value)
        rev = _val(ws.cell(row=totals_row, column=col + 1).value)
        totals.append({'month': MONTH_LABELS[ci], 'expenses': round(exp, 2), 'revenue': round(rev, 2)})

    monthly_net = []
    for ci, col in enumerate(COL_MAP):
        v = _val(ws.cell(row=14, column=col).value)
        monthly_net.append({'month': MONTH_LABELS[ci], 'value': round(v, 2)})

    cumulative = []
    for ci, col in enumerate(COL_MAP):
        v = _val(ws.cell(row=15, column=col).value)
        cumulative.append({'month': MONTH_LABELS[ci], 'value': round(v, 2)})

    wb.close()
    return {
        'month_labels': MONTH_LABELS,
        'projects': projects,
        'totals': totals,
        'monthly_net': monthly_net,
        'cumulative': cumulative,
    }


def load_dashboard_kpis():
    pnl = load_pnl()
    total_revenue = sum(p['summary']['total_revenue'] for p in pnl.values())
    total_expenses = sum(p['summary']['total_op_expenses'] + p['summary']['total_salary_expenses'] for p in pnl.values())
    total_op_profit = sum(p['summary']['total_operating_profit'] for p in pnl.values())
    margin = round((total_op_profit / total_revenue) * 100, 1) if total_revenue > 0 else None

    cf = load_cashflow()
    cash_position = cf['cumulative'][-1]['value'] if cf['cumulative'] else 0

    return {
        'total_revenue': round(total_revenue, 2),
        'total_expenses': round(total_expenses, 2),
        'total_operating_profit': round(total_op_profit, 2),
        'margin': margin,
        'cash_position': round(cash_position, 2),
        'project_summaries': {name: data['summary'] for name, data in pnl.items()},
    }
